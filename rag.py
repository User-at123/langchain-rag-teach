"""RAG 知识库问答：加载文档 → 切分 → 嵌入 → 混合检索（BM25+向量）→ 重排序 → 生成。

支持 txt/md/pdf(含扫描件 OCR)/docx/csv/xlsx；增量索引（.index_state.json）；
含 Text-to-SQL 路由（sql_db.py）。Web 服务入口见 app.py。
"""

import json
import os
import re

import jieba

from dotenv import load_dotenv

# 必须在 import huggingface 相关库之前加载 .env（HF 环境变量只在 import 时读一次）
load_dotenv()

# 国内镜像（可选）：HF_ENDPOINT=https://hf-mirror.com 可加速/修复模型下载
if os.getenv("HF_ENDPOINT"):
    os.environ["HF_ENDPOINT"] = os.getenv("HF_ENDPOINT")

# SSL 证书验证失败时（Windows/公司代理常见），可在 .env 设 USE_INSECURE_SSL=1 跳过验证
if os.getenv("USE_INSECURE_SSL") == "1":
    os.environ["HF_HUB_DISABLE_SSL_VERIFY"] = "1"

from typing import Any, List

from rank_bm25 import BM25Okapi  # 手写 BM25 检索器用（替代已停维护的 langchain-community）

from langchain_chroma import Chroma
from langchain_core.documents import Document
from langchain_core.messages import AIMessage, HumanMessage, SystemMessage
from langchain_core.output_parsers import StrOutputParser
from langchain_core.prompts import ChatPromptTemplate
from langchain_core.retrievers import BaseRetriever
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_openai import ChatOpenAI
from langchain_text_splitters import (
    MarkdownHeaderTextSplitter,
    RecursiveCharacterTextSplitter,
)
from pydantic import ConfigDict, Field, PrivateAttr

from sql_db import SQLiteDb

# ========== 1. 加载文档（支持多格式） ==========
# docs/ 有文件 → 按扩展名路由加载全部；docs/ 不存在或为空 → 回退加载 knowledge_base.txt
DOCS_DIR = os.getenv("DOCS_DIR", "./docs")

# ===== 各格式的加载函数 =====
# 统一签名：fn(path, **kwargs) -> list[Document]，metadata 带 "source" 来源路径
# 文本类必须指定 encoding="utf-8"：Windows 默认 GBK 打开，知识库是 UTF-8，
# 不指定会报 UnicodeDecodeError
def load_text(path, encoding):
    """加载 .txt / .md：直接读入整个文件（手写实现，替代已停维护的 TextLoader）。
    metadata 带 "source" 完整路径——增量索引按 source 删除旧片段。"""
    with open(path, encoding=encoding) as f:
        content = f.read()
    return [Document(page_content=content, metadata={"source": path})]


# ===== PDF 加载（含扫描件自动 OCR）=====
# 判定阈值：一页提取出的文字少于该数量 → 视为扫描页（无文字层，需 OCR）
OCR_MIN_TEXT_LEN = 20
# OCR 结果缓存目录：识别出的文本落盘，下次构建索引直接读缓存，不再重跑 OCR
# 注意：必须在 docs/ 之外（如项目根），否则会被 os.walk 遍历当成文档重复加载
OCR_CACHE_DIR = os.getenv("OCR_CACHE_DIR", "./ocr_cache")


def _ocr_cache_path(path):
    """由 PDF 路径推导缓存文件路径：ocr_cache/<PDF 文件名去扩展名>.txt。"""
    os.makedirs(OCR_CACHE_DIR, exist_ok=True)
    return os.path.join(OCR_CACHE_DIR, os.path.splitext(os.path.basename(path))[0] + ".txt")


def _write_ocr_cache(cache_path, page_texts):
    """把 {页索引: 文本} 写入缓存文件（分隔行格式，正文可含换行）。
    读写都指定 newline="\\n"，避免 Windows 下 \\n 被写成 \\r\\n 导致解析错位。"""
    with open(cache_path, "w", encoding="utf-8", newline="\n") as f:
        for idx in sorted(page_texts):
            f.write(f"<<<PAGE:{idx}>>>\n{page_texts[idx]}\n")


def _read_ocr_cache(cache_path):
    """读缓存文件，返回 {页索引: 文本}。"""
    with open(cache_path, encoding="utf-8", newline="\n") as f:
        content = f.read()
    cached = {}
    for part in content.split("<<<PAGE:")[1:]:
        idx_str, _, text = part.partition(">>>\n")
        cached[int(idx_str)] = text.rstrip("\n")
    return cached


def load_pdf(path):
    """加载 .pdf：文字版直接提取；扫描页（提取文字过少）用 PyMuPDF 渲染 + RapidOCR 离线识别。
    手写文字提取（替代已停维护的 PyPDFLoader）：PyMuPDF 逐页 get_text()，
    输出格式保持"每页一个 Document，metadata 带 source + page"（下方 OCR 判定依赖此结构）。"""
    import pymupdf as fitz

    pdf = fitz.open(path)
    docs = []
    for i in range(len(pdf)):
        docs.append(Document(
            page_content=pdf[i].get_text(),
            metadata={"source": path, "page": i},
        ))
    pdf.close()
    # 找出"扫描页"（提取文字过少）
    ocr_idx = [i for i, d in enumerate(docs)
               if len(d.page_content.strip()) < OCR_MIN_TEXT_LEN]
    if not ocr_idx:
        print(f"  {path}: 文字版 PDF，{len(docs)} 页直接提取（无需 OCR）")
        return docs

    # —— OCR 缓存：识别过的页落盘，命中则直接读，不再重跑 ——
    cache_path = _ocr_cache_path(path)
    if os.path.exists(cache_path) and os.path.getmtime(cache_path) >= os.path.getmtime(path):
        cached = _read_ocr_cache(cache_path)
        hit = 0
        for i in ocr_idx:
            if i in cached:
                docs[i] = Document(
                    page_content=cached[i],
                    metadata={**docs[i].metadata, "ocr": True},
                )
                hit += 1
        print(f"  {path}: 命中 OCR 缓存（{hit}/{len(ocr_idx)} 页），跳过识别，直接读文本")
        return docs

    # 懒加载：只有遇到扫描件才 import/初始化，普通文字版 PDF 不受影响
    print(f"  {path}: {len(ocr_idx)}/{len(docs)} 页是扫描件，开始 OCR（较慢，请耐心）...")
    import pymupdf as fitz  # 渲染 PDF 页为位图（新版包名 pymupdf，兼容旧名 fitz）
    import numpy as np
    from PIL import Image
    from rapidocr_onnxruntime import RapidOCR  # 离线中文 OCR，模型内置
    from tqdm import tqdm  # OCR 进度条

    ocr = RapidOCR()
    pdf = fitz.open(path)
    recognized = {}  # 页索引 -> OCR 文本（跑完一次性写入缓存）
    for i in tqdm(ocr_idx, desc=f"OCR {os.path.basename(path)}", unit="页"):
        pix = pdf[i].get_pixmap(dpi=200)      # 渲染位图（dpi 越高越清晰也越慢）
        img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
        result, _ = ocr(np.array(img))        # [[四角坐标, 文字, 置信度], ...]
        if not result:                        # 识别不出也记空文本并缓存，避免下次重跑
            recognized[i] = ""
            continue
        result.sort(key=lambda r: r[0][0][1])  # 按行从上到下排序，保持阅读顺序
        recognized[i] = "\n".join(r[1] for r in result)
        docs[i] = Document(
            page_content=recognized[i],
            metadata={**docs[i].metadata, "ocr": True},  # 标记 OCR 来源，方便溯源
        )
    pdf.close()
    _write_ocr_cache(cache_path, recognized)
    print(f"  {path}: OCR 完成，共识别 {len(recognized)} 页；结果已缓存到 {cache_path}")
    print(f"          下次构建索引直接读缓存，这本 PDF 不再重跑 OCR")
    return docs


def load_docx(path):
    """加载 .docx：zipfile 解 word/document.xml，按 w:p 段落提取 w:t 文本（手写实现）。
    行为与 Docx2txtLoader 一致：整个文档一个 Document，source 带路径。
    表格内的段落也提取（docx 表格内容同样存在于 w:p 中，与原实现行为一致）。"""
    import zipfile
    from xml.etree import ElementTree as ET

    NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    with zipfile.ZipFile(path) as z:
        xml_content = z.read("word/document.xml")
    root = ET.fromstring(xml_content)
    parts = []
    for p in root.iter(f"{NS}p"):
        text = "".join(t.text or "" for t in p.iter(f"{NS}t"))
        parts.append(text)
    return [Document(page_content="\n".join(parts), metadata={"source": path})]


def load_csv(path, encoding):
    """加载 .csv：每行一个 Document，内容为 "列名: 值" 换行连接（与 LangChain CSVLoader 一致）。
    手写实现：csv.DictReader 读取，空字段/空表头容错（原实现遇 None 值会崩溃）。"""
    import csv

    docs = []
    with open(path, encoding=encoding, newline="") as f:
        for i, row in enumerate(csv.DictReader(f)):
            content = "\n".join(
                f"{k.strip()}: {v.strip()}"
                for k, v in row.items()
                if k is not None and k.strip() and v is not None
            )
            docs.append(Document(
                page_content=content,
                metadata={"source": f"{path}:{i + 2}"},  # 行号从 2 开始（第 1 行是表头）
            ))
    return docs


def load_xlsx(path):
    """加载 .xlsx：第 1 行视为表头，数据行拼成 "字段名: 值 | ..."（行自带字段名）；
    跳过全空行/重复表头行；无表头时整行用 | 连接。"""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    docs = []
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = ["" if c is None else str(c).strip() for c in rows[0]]
        if any(header):  # 有表头：字段名拼进每行数据
            for row_no, row in enumerate(rows[1:], start=2):
                cells = ["" if c is None else str(c).strip() for c in row]
                if not any(cells):
                    continue  # 跳过全空行
                if cells == header:
                    continue  # 跳过与表头重复的行（有些表会重复打印表头）
                pairs = [f"{h}: {v}" for h, v in zip(header, cells) if h and v]
                if not pairs:
                    continue
                docs.append(Document(
                    page_content=" | ".join(pairs),
                    metadata={"source": f"{path} [工作表: {sheet.title} 第{row_no}行]"},
                ))
        else:  # 无表头：整行单元格直接 | 连接（原行为）
            for row_no, row in enumerate(rows, start=1):
                cells = ["" if c is None else str(c) for c in row]
                if not any(cells):
                    continue
                docs.append(Document(
                    page_content=" | ".join(cells),
                    metadata={"source": f"{path} [工作表: {sheet.title} 第{row_no}行]"},
                ))
    return docs


# 扩展名 → (加载函数, 参数) 路由表
LOADER_MAP = {
    ".txt": (load_text, {"encoding": "utf-8"}),
    ".md": (load_text, {"encoding": "utf-8"}),
    ".pdf": (load_pdf, {}),
    ".docx": (load_docx, {}),
    ".csv": (load_csv, {"encoding": "utf-8"}),
    ".xlsx": (load_xlsx, {}),
}


def load_single_file(path):
    """加载单个文件（按扩展名路由）。增量索引的最小单元是一个文件：只加载有变化的那个。"""
    ext = os.path.splitext(path)[1].lower()
    entry = LOADER_MAP.get(ext)
    if entry is None:
        print(f"跳过不支持的文件类型: {os.path.basename(path)}")
        return []
    load_fn, load_kwargs = entry
    return load_fn(path, **load_kwargs)


def load_documents():
    """加载知识库文档，返回 Document 列表（每个带 page_content 和 metadata）。"""
    if os.path.isdir(DOCS_DIR) and any(os.listdir(DOCS_DIR)):
        docs = []
        # 递归遍历 docs/ 下所有文件，逐个加载
        for root, _, files in os.walk(DOCS_DIR):
            for name in files:
                docs.extend(load_single_file(os.path.join(root, name)))
        return docs
    else:
        # 回退：docs/ 为空时加载单文件知识库
        with open("knowledge_base.txt", encoding="utf-8") as f:
            return [Document(page_content=f.read(), metadata={"source": "knowledge_base.txt"})]


# ========== 2. 切分（按格式选择切分器） ==========
# - .md        → 按标题层级切（章节不拆散）
# - .csv/.xlsx → 加载时已按行拆好，不再切分
# - 其他格式   → RecursiveCharacterTextSplitter（chunk_overlap 让相邻块重叠，避免语义切断）
def split_documents_by_format(documents):
    """按文件扩展名路由切分器，返回切分后的 chunks。"""
    chunks = []
    by_ext = {}
    for doc in documents:
        ext = os.path.splitext(doc.metadata.get("source", ""))[1].lower()
        by_ext.setdefault(ext, []).append(doc)

    for ext, docs in by_ext.items():
        if ext == ".md":
            # Markdown：按标题层级切分
            md_splitter = MarkdownHeaderTextSplitter(
                headers_to_split_on=[("#", "一级标题"), ("##", "二级标题"), ("###", "三级标题")],
                strip_headers=False,  # 标题文字保留在正文里，检索时能对上章节
            )
            for doc in docs:
                # 该切分器输入是字符串，切出的块没有 source，需手动补回来源
                for piece in md_splitter.split_text(doc.page_content):
                    piece.metadata["source"] = doc.metadata.get("source", "未知来源")
                    chunks.append(piece)
        elif ext in (".csv", ".xlsx"):
            # 表格：加载时已按行拆好，直接入库，不再切分
            chunks.extend(docs)
        else:
            # 通用切分器；用 split_documents 以保留来源 metadata
            splitter = RecursiveCharacterTextSplitter(chunk_size=200, chunk_overlap=50)
            chunks.extend(splitter.split_documents(docs))
    return chunks


# ========== 3. 嵌入 + 存储（Chroma 持久化 + 增量索引） ==========
# 片段转向量存 ./chroma_db（首次运行自动下载本地模型，约 100MB）
# 注意：DeepSeek 无 Embedding API，故用本地开源模型
# 顶层只留 None 占位，真实对象在 init() 中创建——避免 import 时下载模型/建索引
# （这是 pytest 能跑的前提：测试 import rag 只加载轻量定义，不碰重资源）
embeddings = None
vector_store = None

CHROMA_DIR = os.getenv("CHROMA_DIR", "./chroma_db")

# ===== 增量索引：文件指纹状态管理（.index_state.json） =====
# 记录每个文件的指纹（mtime+size），启动时三向比对：
#   新增 → 只嵌入新文件；变更 → 只重建该文件；删除 → 只删该文件；未变 → 跳过
# 用 mtime+size 而非 md5：快；代价是"内容变了但 mtime/size 恰好没变"会漏检
INDEX_STATE_FILE = os.getenv("INDEX_STATE_FILE", "./.index_state.json")


def file_fingerprint(path):
    """文件指纹：修改时间 + 大小。"""
    st = os.stat(path)
    return f"{st.st_mtime:.3f}-{st.st_size}"


def scan_docs_files():
    """扫描 docs/ 下所有支持的文件，返回 {路径: 指纹}（LOADER_MAP 之外的类型不参与）。"""
    scanned = {}
    if os.path.isdir(DOCS_DIR):
        for root, _, files in os.walk(DOCS_DIR):
            for name in files:
                path = os.path.join(root, name)
                if os.path.splitext(name)[1].lower() in LOADER_MAP:
                    scanned[path] = file_fingerprint(path)
    return scanned


def load_index_state():
    """读 .index_state.json，返回 {路径: 指纹}；文件不存在/解析失败返回 {}。"""
    if os.path.exists(INDEX_STATE_FILE):
        try:
            with open(INDEX_STATE_FILE, encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"  警告：状态文件解析失败（{e}），按无状态处理（全量校准）")
    return {}


def save_index_state(state):
    """把 {路径: 指纹} 写回 .index_state.json（作为下次增量的基线）。"""
    with open(INDEX_STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


# ===== 三向分支：首次全量 / 旧索引校准 / 增量比对 =====
def _build_vector_store():
    """构建或加载向量库（首次全量嵌入，之后增量比对），返回 Chroma 实例。
    逻辑与原顶层代码逐行一致，仅把对全局 vector_store 的赋值改为局部 vs。"""
    scan = scan_docs_files()
    if not os.path.exists(CHROMA_DIR):
        # 分支 1：首次运行——全量加载 + 切分 + 嵌入
        print("首次运行：加载全部文档 ...")
        documents = load_documents()
        print(f"加载了 {len(documents)} 个文档")
        chunks = split_documents_by_format(documents)
        print(f"切分成 {len(chunks)} 个片段")
        print("创建向量索引并保存到磁盘 ...")
        vs = Chroma.from_documents(
            chunks,
            embedding=embeddings,
            persist_directory=CHROMA_DIR,
        )
        save_index_state(scan)  # 全部文件记为"已入库"
        print(f"  已记录 {len(scan)} 个文件的指纹（{INDEX_STATE_FILE}），下次启动开始走增量")
        return vs
    # 索引已存在：直接加载，不重新嵌入
    vs = Chroma(
        embedding_function=embeddings,
        persist_directory=CHROMA_DIR,
    )
    state = load_index_state()
    if not state:
        # 分支 2：有索引但无状态文件——只记指纹，不重新嵌入
        print("检测到已有向量索引但无指纹状态（旧索引）→ 全量校准：记录指纹，不重新嵌入")
        save_index_state(scan)
        print(f"  已记录 {len(scan)} 个文件的指纹，下次启动开始走增量")
    else:
        # 分支 3：增量比对——只处理有变化的文件
        added = [p for p in scan if p not in state]
        changed = [p for p in scan if p in state and scan[p] != state[p]]
        removed = [p for p in state if p not in scan]
        untouched = [p for p in scan if p in state and scan[p] == state[p]]
        print(f"增量比对：新增 {len(added)} 变更 {len(changed)} 删除 {len(removed)} 跳过 {len(untouched)}")
        for p in removed:
            print(f"  [删除] {p}")
            vs.delete(where={"source": p})
        for p in added + changed:
            action = "新增" if p in added else "变更"
            print(f"  [{action}] {p}（单文件处理，其余跳过）")
            docs = load_single_file(p)
            new_chunks = split_documents_by_format(docs)
            if p in changed:
                # 变更：先按 source 删旧片段，再入库新的（避免残留重复）
                vs.delete(where={"source": p})
            vs.add_documents(new_chunks)
        save_index_state(scan)  # 更新基线
        print("  指纹基线已更新")
    return vs

# ========== 4. 创建检索器（混合检索 + 重排序） ==========
# 两路召回（BM25 关键词 + 向量语义）取并集 → 广召回 Top 50 → bge-reranker 精排 → 留 top_n 条
# 注：检索参数变化无需重建 chroma_db（只影响查询，不影响索引内容）

# ===== 3.5 模型（提前定义：MultiQuery 拆子查询与主问答链共用） =====
# 顶层占位，init() 中创建（ChatOpenAI 惰性发请求，但仍统一放 init 保持"import 零副作用"）
llm = None

# ===== 手写工具类：BM25 + RRF 融合 + Reranker 精排 =====
# 官方实现在独立包 langchain-retrievers（国内镜像未同步无法安装），故手写；
# BM25 原 langchain-community 版已停维护，一并改为 rank_bm25 + jieba 手写；
# 依赖只需 rank_bm25 + jieba（均已安装）


class BM25Retriever(BaseRetriever):
    """BM25 关键词检索（手写实现，替代已停维护的 langchain-community BM25Retriever）。

    必须继承 BaseRetriever：LCEL 管道 `| retriever` 和 RRFRetriever.retrievers
    都按 BaseRetriever 协议调用，不继承则不兼容。
    中文分词固定走 jieba（rank_bm25 默认按空格分词，中文整句会被当成一个 token）。"""

    model_config = ConfigDict(arbitrary_types_allowed=True)

    documents: List[str] = Field(description="语料文本列表")
    metadatas: List[dict] = Field(default_factory=list, description="与 documents 对应的元数据")
    k: int = Field(default=50, description="返回条数")

    _bm25: Any = PrivateAttr(default=None)

    def _init_bm25(self):
        """懒初始化：jieba 分词后建 BM25 索引（与构建时刻分离，import 零开销）。"""
        self._bm25 = BM25Okapi([list(jieba.cut(t)) for t in self.documents])

    def _get_relevant_documents(self, query: str, *, run_manager: Any = None):
        if self._bm25 is None:
            self._init_bm25()
        tokens = list(jieba.cut(query))
        if not tokens:
            return []
        scores = self._bm25.get_scores(tokens)
        top_idx = sorted(range(len(scores)), key=lambda i: scores[i], reverse=True)[: self.k]
        return [
            Document(
                page_content=self.documents[i],
                metadata=self.metadatas[i] if i < len(self.metadatas) else {},
            )
            for i in top_idx
        ]

    @classmethod
    def from_texts(cls, texts, k=50, preprocess_func=None, metadatas=None):
        """兼容原 LangChain 版调用签名（preprocess_func 保留参数但不使用，分词内部固定 jieba）。"""
        return cls(documents=list(texts), k=k, metadatas=list(metadatas or []))


class RRFRetriever(BaseRetriever):
    """互惠排名融合（RRF）：多路 Top-k 排名按 1/(k+rank) 融合，只比排名不比分数。"""

    retrievers: List[BaseRetriever] = Field(description="多路检索器（如 BM25 + 向量）")
    rrf_k: int = Field(default=60, description="RRF 常数（经验值 60）")
    top_n: int = Field(default=50, description="融合后保留的候选条数")

    def _get_relevant_documents(self, query: str, *, run_manager: Any = None):
        scores: dict = {}    # page_content -> 累计融合分
        docs_map: dict = {}  # page_content -> Document（保留 metadata）
        for retriever in self.retrievers:
            for rank, doc in enumerate(retriever.invoke(query)):
                key = doc.page_content
                docs_map.setdefault(key, doc)
                scores[key] = scores.get(key, 0.0) + 1.0 / (self.rrf_k + rank)
        ranked = sorted(
            docs_map.values(),
            key=lambda d: scores[d.page_content],
            reverse=True,
        )
        return ranked[: self.top_n]


class RerankerRetriever(BaseRetriever):
    """bge-reranker 精排：粗召回候选 → 交叉编码器逐条算相关度 → 重排留 top_n 条。"""

    model_config = ConfigDict(arbitrary_types_allowed=True)

    base_retriever: BaseRetriever = Field(description="粗召回检索器（混合检索）")
    model_name: str = Field(
        default="BAAI/bge-reranker-base",
        description="交叉编码器模型名（首次运行自动下载）",
    )
    top_n: int = Field(
        default=6,
        description="精排后保留的条数（最终进提示词的视野，承接 V7.7 的 k=6 经验）",
    )

    _cross_encoder: Any = PrivateAttr(default=None)

    def _get_cross_encoder(self):
        """懒加载交叉编码器：首次调用才下载模型，避免 import 阶段联网卡住"""
        if self._cross_encoder is None:
            from sentence_transformers import CrossEncoder

            self._cross_encoder = CrossEncoder(self.model_name)
        return self._cross_encoder

    def _get_relevant_documents(self, query: str, *, run_manager: Any = None):
        candidates = self.base_retriever.invoke(query)     # 粗召回 Top 50
        pairs = [(query, d.page_content) for d in candidates]
        scores = self._get_cross_encoder().predict(pairs)  # 逐条相关度，越大越相关
        ranked = sorted(zip(candidates, scores), key=lambda x: x[1], reverse=True)
        return [d for d, _ in ranked[: self.top_n]]


class MultiQueryRetriever(BaseRetriever):
    """多路子查询检索：LLM 把问题拆成多个子查询（覆盖同一事实的不同说法，
    解决"枚举/聚合"类问题单路检索漏全集）。

    三步：拆子查询 → 每路粗召回后合并成一个大 batch 交给 reranker 精排 →
    每路取精排前 per_query_top_n，按 rank 融合分排序去重 → 最终 top_n。
    """

    model_config = ConfigDict(arbitrary_types_allowed=True)

    base_retriever: BaseRetriever = Field(description="粗召回检索器（RRF 混合检索）")
    llm: Any = Field(description="生成子查询的 LLM（一次调用生成多路）")
    sub_query_count: int = Field(default=5, description="拆成几路子查询")
    per_query_top_n: int = Field(default=5, description="每路子查询精排后保留条数（给弱相关但关键的信息留空间）")
    recall_top_n: int = Field(
        default=30,
        description="每路子查询粗召回条数。V11.1 实测：抽象问法（结局/下场）改写出的"
        "「田雨 自杀」单路在 RRF 全量排第 23，recall_top_n=20 会把它截断丢关键信息；"
        "提到 30 兜住罕见词 chunk。30 条/路 × 5 路去重后 ≈ 150 对精排 ≈ 15s（本地实测），"
        "与旧版 20 条/路耗时相当（瓶颈在 reranker 首轮，去重后对数量增长有限）。",
    )
    top_n: int = Field(default=6, description="最终进 LLM 视野的条数")
    model_name: str = Field(
        default="BAAI/bge-reranker-base",
        description="交叉编码器模型名（与 RerankerRetriever 同款）",
    )

    _cross_encoder: Any = PrivateAttr(default=None)

    def _get_cross_encoder(self):
        """懒加载交叉编码器：首次调用才下载模型（与 RerankerRetriever 同款）"""
        if self._cross_encoder is None:
            from sentence_transformers import CrossEncoder

            self._cross_encoder = CrossEncoder(self.model_name)
        return self._cross_encoder

    def _generate_sub_queries(self, query):
        """一次 LLM 调用生成子查询列表（JSON 数组）。
        解析失败逐级降级：json.loads → 正则抓 [...] → 返回原问题本身。"""
        import json
        import re

        prompt = (
            f"你是检索辅助器。请把用户问题改写成 {self.sub_query_count} 个不同的子查询，"
            "用于分别检索知识库。\n"
            "要求：\n"
            "1. 如果问题在问一个\"集合\"（如\"有几个/都有谁/分别是\"），请为每个可能的成员"
            "单独生成一个子查询，并直接带上该成员的名字或专有名词"
            "（如\"李云龙第一任妻子 杨秀芹\"、\"李云龙第二任妻子 田雨\"）；\n"
            "2. 每个子查询只聚焦一个对象/一个说法，不要在一个子查询里混多个名字；\n"
            "3. 可补充\"关系式\"问法（如\"杨秀芹和李云龙是什么关系\"），绕开原问题里的关键词；\n"
            "4. 抽象问法必须具象化（做\"方向拆解\"）：用户可能用抽象/概括的说法询问（如某人的"
            "结局/下场/命运/后来如何、内心情感/心情、一段关系/爱情的发展、某人的去向/下落、"
            "某段际遇/事业/财产/战斗结果等）。对这类说法，请像一位熟悉作品的读者那样，"
            "预想原文更可能用哪些具体词/情节/情境来表达这个抽象概念（不局限于以下示例，"
            "凡抽象概括处都应如此预想）：\n"
            "   - 结局/命运 → 死亡/自杀/牺牲/去世/被捕/归隐/下落不明…\n"
            "   - 心情/情绪 → 愤怒/悲痛/得意/沉默/流泪/狂喜…\n"
            "   - 爱情/关系 → 结婚/表白/离婚/分手/私奔/反目…\n"
            "   - 去向/下落 → 被俘/阵亡/逃亡/回家/奔赴…\n"
            "   - 婚姻/事业/战斗结果 → 结婚/离婚 或 升迁/贬职 或 胜利/战败…\n"
            "然后把这些预想的具体词分散到不同子查询，每个方向各写一路，不要所有子查询都共用"
            "同一个具体词——方向随问题自由预想，不局限于本示例；\n"
            "5. 死亡/生命状态类是方向拆解的重要案例（易漏高价值）：对结局/下场/命运/后来怎么样等"
            "死亡类抽象词，必须同时覆盖正常死亡（去世/病故/终老）与非正常死亡（自杀/牺牲/被杀/"
            "处决/自尽）两个方向，各成一路——答案往往藏在非正常死亡的原文写法里，"
            "只联想\"去世\"这类通用词命不中；\n"
            "6. 至少保留一路子查询直接使用原问题的原始关键词（防止改写丢失原文信息）；\n"
            "7. 专有名词（人名/地名/机构名/作品名）必须原样带上，不得省略或改写；\n"
            "8. 每个子查询包含足够上下文词，能独立完成检索；\n"
            "9. 只返回 JSON 数组字符串（如 [\"子查询1\", \"子查询2\", ...]），不要任何其他文字。\n\n"
            f"用户问题：{query}"
        )
        resp = self.llm.invoke(prompt)
        text = getattr(resp, "content", str(resp)).strip()

        def _try_parse(t):
            arr = json.loads(t)
            if isinstance(arr, list):
                return [s.strip() for s in arr if isinstance(s, str) and s.strip()]
            return None

        try:
            subs = _try_parse(text)
            if subs:
                return subs[: self.sub_query_count]
        except Exception:
            pass
        m = re.search(r"\[.*\]", text, re.S)
        if m:
            try:
                subs = _try_parse(m.group(0))
                if subs:
                    return subs[: self.sub_query_count]
            except Exception:
                pass
        return [query]  # 解析失败：退化用原问题本身

    def _get_relevant_documents(self, query: str, *, run_manager: Any = None):
        sub_queries = self._generate_sub_queries(query)
        # 并行粗召回（V10.3 性能优化）：多路子查询互相独立，串行会白白累加等待时间。
        # 第一个子查询在主线程先跑——顺带触发 BM25 懒索引构建（jieba 分词 2541 片段），
        # 避免多线程同时首调导致重复构建 + GIL 争抢；其余子查询交线程池并行
        # （本地 BM25/Chroma 在 native 调用处会释放 GIL，实际省大部分串行时间）。
        from concurrent.futures import ThreadPoolExecutor

        route_results = {}
        if sub_queries:
            first = sub_queries[0]
            try:
                route_results[first] = self.base_retriever.invoke(first)[: self.recall_top_n]
            except Exception as e:
                print(f"  [检索] 子查询「{first}」失败，忽略：{e}")
                route_results[first] = []
        with ThreadPoolExecutor(max_workers=min(4, max(1, len(sub_queries) - 1))) as pool:
            futures = {
                pool.submit(self.base_retriever.invoke, sq): sq
                for sq in sub_queries[1:]
            }
            for fut, sq in futures.items():
                try:
                    route_results[sq] = fut.result()[: self.recall_top_n]
                except Exception as e:
                    print(f"  [检索] 子查询「{sq}」失败，忽略：{e}")
                    route_results[sq] = []
        # 每路粗召回，收集 (子查询, Document) 对；跨路去重：同一 page_content 只保留一次
        collected = []
        seen = set()
        raw_count = 0
        for sq in sub_queries:
            for doc in route_results.get(sq, []):
                raw_count += 1
                if doc.page_content in seen:
                    continue  # 已被其他子查询召回，跳过
                seen.add(doc.page_content)
                collected.append((sq, doc))
        print(f"  跨路去重：{raw_count} 对候选 → {len(collected)} 对唯一片段（节省 {raw_count - len(collected)}）")
        # 提速：全部候选对合并成一个 batch 一次 predict（模型只加载一次）
        # 显式标注类型，避免类型检查把 list 推断成 tuple
        pairs: List[List[str]] = [[sq, d.page_content] for sq, d in collected]
        scores = self._get_cross_encoder().predict(pairs, batch_size=64)
        # 按子查询分组，每路取精排前 per_query_top_n，rank 融合排序
        by_route = {}
        for (sq, doc), s in zip(collected, scores):
            by_route.setdefault(sq, []).append((doc, s))
        fusion = {}  # page_content -> [Document, 融合分]
        for sq, items in by_route.items():
            for rank, (doc, s) in enumerate(
                sorted(items, key=lambda x: x[1], reverse=True)
            ):
                if rank >= self.per_query_top_n:
                    break  # 每路只保留精排前 per_query_top_n 条
                key = doc.page_content
                if key not in fusion:
                    fusion[key] = [doc, 0.0]
                # 只比排名不比分数：不同子查询打出的分数量纲不可比（实验已验证）
                fusion[key][1] += 1.0 / (60 + rank)
        ranked = sorted(fusion.values(), key=lambda x: x[1], reverse=True)
        return [d for d, _ in ranked[: self.top_n]]


# 4.x 检索器构建（init() 中调用；依赖 vector_store 与 llm）
# 注：bm25/vector/ensemble/reranker 检索器仅在构建时组合使用，无需模块级暴露
def _build_retrievers(vs, llm_obj):
    """基于向量库与 LLM 构建混合检索器，返回最终 MultiQueryRetriever（进管道的那一个）。"""
    # 4.1 BM25 关键词检索（无状态：每次启动从向量库全量取回文本现场重建）
    #     坑：rank_bm25 默认按空格分词，中文整句会被当成一个 token，BM25 失效；
    #         必须传 jieba 分词器（preprocess_func；新版独立包叫 tokenizer）
    _all_docs = vs.get(include=["documents", "metadatas"])
    bm25_texts = _all_docs["documents"]
    bm25_metadatas = _all_docs["metadatas"]
    print(f"BM25 从向量库取回 {len(bm25_texts)} 个片段重建（无状态索引）")
    bm25_retriever = BM25Retriever.from_texts(
        bm25_texts,
        k=50,
        preprocess_func=lambda t: list(jieba.cut(t)),
        metadatas=bm25_metadatas,  # 保留 source 等元数据，方便追溯
    )

    # 4.2 向量语义检索（Chroma，召回放宽到 50，交由 reranker 精排）
    vector_retriever = vs.as_retriever(search_kwargs={"k": 50})

    # 4.3 融合两路召回（手写 RRF）
    ensemble_retriever = RRFRetriever(
        retrievers=[bm25_retriever, vector_retriever],
    )

    # 4.4 重排序（交叉编码器：候选逐条算相关度，精排后只留 top_n 条）
    # 注意：以下实例未进当前管道（管道用 4.5 的 MultiQuery，内部自带批量精排），
    #       保留作单路对比演示；删掉不影响运行
    reranker_retriever = RerankerRetriever(
        base_retriever=ensemble_retriever,
        model_name=os.getenv("RERANKER_MODEL", "BAAI/bge-reranker-base"),
        top_n=6,
    )

    # 4.5 MultiQuery 多路检索：LLM 拆子查询覆盖"不同说法"→ 每路独立召回精排 → rank 融合。
    #     解决"枚举/聚合"类问题（几个/都有谁）单路检索漏全集。
    #     管道 `| retriever |` 不用改，换的只是 retriever 实现。
    return MultiQueryRetriever(
        base_retriever=ensemble_retriever,
        llm=llm_obj,
        sub_query_count=5,
        per_query_top_n=5,  # V11.1：3→5，精排后每路多留 2 条，给"弱相关但关键"的罕见词 chunk 留空间
        recall_top_n=30,  # V11.1：15→30。实测「田雨 结局 自杀」单路 RRF 排第 25，15 会截断丢答案；
        # 30 条/路 × 5 路去重后约 150 对精排 ≈ 15s（V10.3 曾砍到 15 省 ~2s，但牺牲了召回，回调）
        top_n=8,  # V11.1.3：6→8。实测「田雨死了吗」割腕 chunk 精排第 2（分 0.94）但只在一路出现，
        # RRF 融合分 1/62 排第 7——top6 被"多路同现"的常规 chunk 占满，关键 chunk 卡在第 7 被截断。
        # 8 条视野对"只在一路出现但高度相关"的 chunk 更宽容；LLM 上下文仅多 ~1k 字（deepseek 可承受）
    )

# ========== 5. 提示词 ==========
# 检索结果作为"参考资料"进提示词，让模型据此回答
prompt = ChatPromptTemplate.from_messages([
    ("system",
     "你是一个助手。请优先根据【参考资料】回答问题；"
     "如果参考资料里没有答案，请如实说明。\n\n"
     "【参考资料】\n{context}"),
    ("human", "{question}"),
])

# ========== 6. 组合成链 ==========
def format_docs(docs):
    """把检索到的文档列表拼成一段干净文本，并标注来源文件名。"""
    parts = []
    for doc in docs:
        source = doc.metadata.get("source", "未知来源")
        parts.append(f"[来源: {source}]\n{doc.page_content}")
    return "\n\n".join(parts)

# 并行分支：context = 检索 question 并格式化成文本；question = 原样透传
def build_qa_chain(llm_obj, retriever_obj):
    """构建主问答链：检索 question → 格式化 context → 提示词 → LLM。
    独立成函数，便于 init() 构建全局 chain、测试时注入 mock llm/retriever 重建。"""
    return (
        {
            "context": (lambda x: x["question"]) | retriever_obj | format_docs,
            "question": lambda x: x["question"],
        }
        | prompt
        | llm_obj
    )


chain = None  # init() 中构建
retriever = None  # init() 中构建（MultiQueryRetriever 实例）

# ========== 6.5 Text-to-SQL + 路由器 ==========
# 链路：路由器判断类型 → 事实查询走向量链（chain）→ 列举/统计走 SQL 链（sql_db.py）；
#       SQL 执行失败 → 报错回传 LLM 重写 1 次；空结果/仍失败 → 降级向量链
# 术语翻译靠 LLM 常识（schema 只给真实表头+样例），捏造列名由 sql_db 列名校验拦截
db = None          # init() 中创建（SQLiteDb 实例）
router_chain = None  # init() 中构建（router_prompt | llm | StrOutputParser）

# 6.5.1 路由器：判断问题走 vector（文档检索）/ sql（数据库查询）/ memory（结合对话历史）
#   / agent（多跳任务自主编排）
# V11.0（第 9 步）新增 memory 路：历史上下文只在判断时提供给 LLM，不参与 vector/sql 的回答
# V12.0（第 10 步）新增 agent 路：多步/组合问题交给 Agent 循环（三个工具自主编排），
#   路由器从"唯一决策者"降级为"调度员"——单跳走快路径，多跳进 Agent，Agent 失败回退快路径
router_prompt = ChatPromptTemplate.from_messages([
    ("system",
     "你是问题路由器。判断用户问题应该走哪条查询路径，只返回一个词：vector、sql、memory 或 agent。\n\n"
     "- vector：答案在叙述性文档里（问属性/关系/定义/背景），读几段原文就能答；\n"
     "  例如\"华信银行的行业是什么\"、\"李云龙的妻子是谁\"。\n"
     "- sql：需要对表格数据做列举/统计/聚合/排名，必须查数据库才能答；\n"
     "  例如\"所有客户的行业有哪些\"、\"客户总数是多少\"、\"各行业客户数量排名\"。\n"
     "- memory：问题引用了本对话之前的内容（指代、追问，或\"刚才/上轮/之前/第 X 条\"），\n"
     "  只有结合对话历史才能理解；例如\"那她们结局如何\"、\"刚才结果里金融客户有谁\"。\n"
     "- agent：需要多步/组合才能完成（先查再分析、结合文档与表格、先统计再总结），\n"
     "  单条路径的固定链路不够用；例如\"先统计各行业客户数，再针对最多的行业生成一句话介绍\"。\n\n"
     "判断规则：\n"
     "1. 含多步/组合信号（先…再/然后/接着/结合…与…/分别…并…/既…又…）→ agent\n"
     "2. 出现统计/聚合词（多少/几个/总数/合计/平均/最大/最小/排行/排名/占比）→ sql\n"
     "3. 出现全集列举词（所有/全部/每个/各自/名单/明细/清单）→ sql\n"
     "4. 问题含指代（她们/他们/它/这个/那个/刚才/上轮/之前/结果/第X条/那...）\n"
     "   且下方对话历史非空 → memory\n"
     "5. 其余（是什么/怎么样/为什么/谁/关系/约定/条款）→ vector\n"
     "6. 不确定时默认 vector（文档检索覆盖面更广，SQL 只覆盖表格数据）\n\n"
     "对话历史（仅用于判断是否引用上轮，不要用它回答问题）：\n{history_hint}"),
    ("human", "{question}"),
])
# router_chain 在 init() 中构建（依赖 llm）：
#   router_chain = router_prompt | llm | StrOutputParser()

# 6.5.2 Text-to-SQL：问题 + schema（真实表头/样例）→ LLM 生成 SELECT
sql_gen_prompt = ChatPromptTemplate.from_messages([
    ("system",
     "你是 SQL 生成器。根据表结构把用户问题转成一条 SQLite SELECT 语句。\n\n"
     "表结构（列名可能与用户说法不同，请按语义匹配最合适的列，\n"
     "比如用户说\"诞生时间/成立时间\"，表里列名可能是\"创立日期\"）：\n{schema}\n\n"
     "要求：\n"
     "- 只输出 SQL 语句本身，不要任何解释\n"
     "- 只允许 SELECT，禁止 INSERT/UPDATE/DELETE/DROP\n"
     "- 中文值要加单引号，表名/列名不要加引号\n"
     "- 如果问题与表结构完全无关，只输出：NULL\n\n"
     "示例：\n"
     "问：所有客户的行业有哪些？\n"
     "答：SELECT DISTINCT 行业 FROM customers\n"
     "问：客户总数是多少？\n"
     "答：SELECT COUNT(*) FROM customers"),
    ("human", "{question}"),
])

# 6.5.3 SQL 执行失败时的重写提示词（报错回传，最多重试 1 次）
sql_retry_prompt = ChatPromptTemplate.from_messages([
    ("system",
     "你生成的 SQL 执行失败。请根据错误信息和表结构重新生成一条 SQLite SELECT 语句。\n\n"
     "表结构：\n{schema}\n"
     "你上次的 SQL：{sql}\n"
     "错误信息：{error}\n\n"
     "只输出修正后的 SQL 语句本身，不要解释。"),
    ("human", "{question}"),
])


def format_sql_result(cols, rows):
    """SQL 结果格式化成文本（纯拼接，不额外调 LLM）。
    注意：rows 是 dict 列表，取值必须 r[c]（zip 迭代 dict 取到的是键不是值）。"""
    parts = [f"[数据库查询] 共 {len(rows)} 条结果："]
    for r in rows:
        parts.append("  " + " | ".join(f"{c}: {r[c]}" for c in cols))
    return "\n".join(parts)


def sql_answer(question, llm_obj, db_obj):
    """SQL 链：生成 SQL → 执行 → 报错重写 1 次；空结果/失败返回 None 由门面降级。
    llm_obj / db_obj 由调用方注入（默认走全局，测试可传 mock，省真实 API 调用）。"""
    schema = db_obj.schema_text()
    sql = (sql_gen_prompt | llm_obj | StrOutputParser()).invoke(
        {"question": question, "schema": schema}
    ).strip()
    print(f"  [SQL 链] 生成的 SQL：{sql}")
    if sql.upper() == "NULL":
        return None  # LLM 判断与表无关
    try:
        cols, rows = db_obj.query(sql)
    except Exception as e:
        retry = (sql_retry_prompt | llm_obj | StrOutputParser()).invoke(
            {"question": question, "schema": schema, "sql": sql, "error": str(e)}
        ).strip()
        print(f"  [SQL 链] 重写：{retry}")
        try:
            cols, rows = db_obj.query(retry)
        except Exception as e2:
            print(f"  [SQL 链] 重写后仍失败，降级向量检索：{e2}")
            return None
    if not rows:
        return None  # 空结果：数据可能不在表里 → 交给向量链
    return format_sql_result(cols, rows)


# ========== 6.7 多轮记忆 + 跨轮状态（第 9 步） ==========
# 会话存储：内存 dict（本地单进程够用；与回答缓存同生命周期，重启即失）
# 每个 session_id 维护：
#   history: [(role, text), ...]——最近 MAX_HISTORY_ROUNDS 轮对话（补全器的原料）
#   last_result: 最近一次 SQL 查询结果的文本（跨轮状态：追问"刚才结果里..."直接取用，不重查库）
sessions: dict = {}
MAX_HISTORY_ROUNDS = 6


def _session(session_id):
    return sessions.setdefault(session_id, {"history": [], "last_result": None})


# 6.7.1 记忆补全器：把"引用上轮"的问题改写成自包含问题，并判断重写后走哪条路
memory_prompt = ChatPromptTemplate.from_messages([
    ("system",
     "你是对话记忆处理器。用户在多轮对话中，当前问题可能引用之前的对话内容"
     "（指代词如\"她们/它们/它/这个/那个/刚才/上轮/第 X 条\"）。\n\n"
     "最近对话历史：\n{history}\n\n"
     "上轮查询结果（若本轮直接引用它，这里会有内容）：\n{last_result}\n\n"
     "请把当前问题改写成脱离上下文也能独立理解的完整问题，并判断重写后的问题"
     "应该怎么回答。只返回一行 JSON，不要任何其他文字，格式为：\n"
     '{{"rewritten": "重写后的完整问题", "base": "vector" 或 "sql" 或 "last_result"}}\n\n'
     "- base=vector：答案在文档资料里（问属性/关系/背景），重写后走文档检索；\n"
     "- base=sql：需要查数据库做列举/统计/聚合，重写后走数据库查询；\n"
     "- base=last_result：上轮查询结果里直接有答案（如\"刚才的结果\"、\"第 X 条\"），\n"
     "  此时 rewritten 给原问题即可，不用重写。\n\n"
     "示例：\n"
     "历史：user：李云龙的妻子是谁？assistant：李云龙的两位妻子是杨秀芹和田雨。\n"
     "当前问题：那她们结局如何？\n"
     '输出：{{"rewritten": "李云龙的两位妻子杨秀芹和田雨的结局如何？", "base": "vector"}}'),
    ("human", "{question}"),
])

# 6.7.2 基于上轮结果直接回答（base=last_result，不再检索/查库）
qa_from_result_prompt = ChatPromptTemplate.from_messages([
    ("system",
     "你是助手。用户引用了上一轮的查询结果，请直接基于下面的【上轮结果】回答，"
     "不要虚构结果里没有的信息。\n\n"
     "【上轮结果】\n{result}"),
    ("human", "{question}"),
])


def _parse_memory_plan(text):
    """解析补全器输出的 JSON {rewritten, base}；失败返回 None（调用方降级走普通路由）。"""
    import json
    import re

    def _try(t):
        data = json.loads(t)
        if not isinstance(data, dict):
            return None
        rewritten = str(data.get("rewritten", "")).strip()
        base = data.get("base", "vector")
        if not rewritten:
            return None
        if base not in ("vector", "sql", "last_result"):
            base = "vector"
        return rewritten, base

    try:
        plan = _try(text)
        if plan:
            return plan
    except Exception:
        pass
    m = re.search(r"\{.*\}", text, re.S)  # 容忍模型输出前后夹杂说明文字
    if m:
        try:
            plan = _try(m.group(0))
            if plan:
                return plan
        except Exception:
            pass
    return None


def _answer_with_memory(question, session, _llm, _db, _chain, _router):
    """memory 路由：补全器改写问题 → 按 base 分流（vector / sql / last_result）。"""
    pairs = session["history"][-MAX_HISTORY_ROUNDS * 2:]
    history = "\n".join(f"{role}：{text}" for role, text in pairs) or "（无历史）"
    last_result = session.get("last_result") or "（无）"
    resp = (memory_prompt | _llm | StrOutputParser()).invoke({
        "question": question,
        "history": history,
        "last_result": last_result,
    }).strip()
    print(f"  [记忆] 补全器输出：{resp}")
    plan = _parse_memory_plan(resp)
    if plan is None:
        print("  [记忆] 补全器输出无法解析，按普通问题走完整路由")
        return _run_route(question, _llm, _db, _chain, _router, history_hint=history)
    rewritten, base = plan
    if base == "last_result" and session.get("last_result"):
        print("  [记忆] 直接基于上轮结果回答（不检索、不查库）")
        return (qa_from_result_prompt | _llm | StrOutputParser()).invoke({
            "question": question,
            "result": session["last_result"],
        })
    if base == "sql":
        ans = sql_answer(rewritten, _llm, _db)
        if ans is not None:
            return ans
        print("  [记忆] SQL 链无结果，降级向量检索")
    return _chain.invoke({"question": rewritten}).content


def _run_route(question, _llm, _db, _chain, _router, history_hint=""):
    """单次 vector/sql 分流并返回回答文本（memory 路不在此处理，由 ask 外层接管）。"""
    route = _router.invoke({"question": question, "history_hint": history_hint}).strip()
    print(f"  [路由器] 问题类型：{route}")
    if route == "sql":
        ans = sql_answer(question, _llm, _db)
        if ans is not None:
            return ans
        print("  [路由器] SQL 链无结果，降级向量检索")
    return _chain.invoke({"question": question}).content


# ========== 6.8 Agent 工具化（第 10 步） ==========
# 手写 ReAct 循环（零新增依赖，不引 langgraph）：把 query_sql / retrieve_vector /
# read_memory 封装成工具，LLM 在 Thought→Action→Observation 循环里自主编排多步。
# 路由器负责"什么时候进 Agent"（多跳信号），Agent 负责"多步怎么走"。
MAX_AGENT_STEPS = 4  # 防死循环：超过步数仍未输出 Final Answer 就回退快路径

AGENT_SYSTEM = (
    "你是自主 Agent，负责拆解需要多步才能完成的任务。你有三个工具，按需调用、可多次组合：\n\n"
    "- query_sql(问题)：把问题转成 SQLite SELECT 语句并执行，返回结果文本。\n"
    "  适合统计/列举/聚合（多少/几个/总数/排名/有哪些）。\n"
    "- retrieve_vector(问题)：检索知识库文档并生成回答。\n"
    "  适合属性/关系/背景类事实问题（是什么/谁/为什么）。\n"
    "- read_memory()：读取本会话的历史对话与上轮查询结果。\n"
    "  追问/引用上轮内容时先用它取上下文。\n\n"
    "每轮输出必须是以下两种格式之一（不要输出其他文字）：\n"
    "Action: 工具名\n"
    "Action Input: 参数\n\n"
    "或者任务完成时输出：\n"
    "Final Answer: 最终回答\n\n"
    "规则：\n"
    "1. 先分析任务分几步，每步选最合适的工具；\n"
    "2. 工具结果会作为 Observation 返回给你，看到 Observation 再决定下一步；\n"
    "3. 任务完成立刻输出 Final Answer，不要多余动作；\n"
    "4. 工具都无法解决时，在 Final Answer 里如实说明。"
)


def _extract_final(text):
    """提取 Final Answer 后面的内容；没有则返回 None。"""
    m = re.search(r"(?:Final Answer|最终答案)\s*[：:]\s*(.+)", text, re.S)
    return m.group(1).strip() if m else None


def _parse_action(text):
    """解析 Action / Action Input 两行；缺 Action 返回 (None, None)。"""
    am = re.search(r"Action\s*[：:]\s*(\w+)", text)
    im = re.search(r"Action Input\s*[：:]\s*(.+)", text, re.S)
    if not am:
        return None, None
    name = am.group(1).strip()
    arg = im.group(1).strip() if im else ""
    return name, arg


def _format_session_context(session):
    """read_memory 工具：把会话历史 + 上轮结果拼成给 Agent 的上下文文本。"""
    if session is None:
        return "（无会话上下文）"
    history = "\n".join(
        f"{role}：{text}" for role, text in session["history"][-MAX_HISTORY_ROUNDS * 2:]
    ) or "（无历史）"
    last = session.get("last_result") or "（无）"
    return f"对话历史：\n{history}\n\n上轮查询结果：\n{last}"


def _call_agent_tool(name, arg, _llm, _db, _chain, session):
    """调度三个 Agent 工具，返回 Observation 文本。"""
    if name == "query_sql":
        if _db is None:
            return "（数据库不可用，无法执行查询）"  # 注入模式未传 db 时不崩，由 Agent/门面降级
        ans = sql_answer(arg, _llm, _db)
        if ans is not None and session is not None:
            session["last_result"] = ans  # 同步跨轮状态，Agent 后续 read_memory 能读到本轮结果
        return ans if ans is not None else "（SQL 查询无结果：该问题可能不在数据库表格数据中）"
    if name == "retrieve_vector":
        return _chain.invoke({"question": arg}).content
    if name == "read_memory":
        return _format_session_context(session)
    return f"（未知工具 {name}，可用：query_sql / retrieve_vector / read_memory）"


def _run_agent(question, _llm, _db, _chain, session):
    """手写 ReAct 循环：Thought→Action→Observation 直到 Final Answer 或超步数。

    每轮把完整对话（SystemMessage + 历史消息 + 新 Observation）直接传给 LLM，
    循环由 LLM 的"看到 Observation 再决策"驱动；超步数或输出不可解析返回 None，
    由 ask 门面回退快路径（_run_route），保证 Agent 异常不拖垮整个问答。
    """
    messages = [SystemMessage(content=AGENT_SYSTEM), HumanMessage(content=question)]
    for step in range(1, MAX_AGENT_STEPS + 1):
        resp = _llm(messages)  # __call__ 等价 invoke：真实 ChatOpenAI 与测试裸 callable 都兼容
        text = resp.content if hasattr(resp, "content") else str(resp)
        print(f"  [Agent 第 {step} 步] {text}")
        final = _extract_final(text)
        if final is not None:
            return final
        action, arg = _parse_action(text)
        if action is None:
            print("  [Agent] 输出无法解析，回退快路径")
            return None
        observation = _call_agent_tool(action, arg, _llm, _db, _chain, session)
        print(f"  [Agent] 工具 {action} 观察结果：{observation[:80]}...")
        messages.append(AIMessage(content=text))
        messages.append(HumanMessage(content=f"Observation: {observation}\n看到结果后请继续（下一步 Action 或 Final Answer）"))
    print(f"  [Agent] 超过 {MAX_AGENT_STEPS} 步未收敛，回退快路径")
    return None


def _record_session(session, question, response):
    """把一轮问答写入会话（历史截断到 MAX_HISTORY_ROUNDS 轮）；SQL 结果暂存为跨轮状态。"""
    session["history"].append(("user", question))
    session["history"].append(("assistant", response))
    if len(session["history"]) > MAX_HISTORY_ROUNDS * 2:
        del session["history"][: len(session["history"]) - MAX_HISTORY_ROUNDS * 2]
    if response.startswith("[数据库查询]"):
        session["last_result"] = response


def init():
    """初始化所有重资源（幂等）：嵌入模型 → 向量库（首次全量/增量）→ 检索器 → LLM → SQLite → 链。

    import rag 不自动执行本函数（顶层零副作用）；调用时机：
    - app.py 启动时（lifespan）调用一次；
    - rag.py CLI 入口进入前调用；
    - 测试 import rag 后可跳过（测纯函数不需要这些对象）。
    """
    global embeddings, vector_store, llm, db, retriever, chain, router_chain
    if embeddings is None:
        embeddings = HuggingFaceEmbeddings(
            model_name=os.getenv("EMBEDDING_MODEL", "BAAI/bge-small-zh-v1.5"),
        )
    if vector_store is None:
        vector_store = _build_vector_store()
    if llm is None:
        llm = ChatOpenAI(
            base_url=os.getenv("OPENAI_BASE_URL", "https://api.deepseek.com/v1"),
            model=os.getenv("LLM_MODEL", "deepseek-chat"),
            temperature=0.3,
        )
    if db is None:
        db = SQLiteDb()
    if retriever is None:
        retriever = _build_retrievers(vector_store, llm)
    if chain is None:
        chain = build_qa_chain(llm, retriever)
    if router_chain is None:
        router_chain = router_prompt | llm | StrOutputParser()


_answer_cache: dict = {}  # V10.3：相同问题直接命中缓存秒回（本地运行，内存足够）
_MAX_CACHE_SIZE = 64


def ask(question, *, session_id=None, llm=None, retriever=None, db=None):
    """门面：单一入口。四路路由器分流 vector / sql / memory / agent（第 10 步 Agent 工具化）。

    - vector：文档检索链；sql：Text-to-SQL 链（失败自动降级向量链）；
    - memory：结合本会话历史回答——补全指代后重路由，或直接引用上轮 SQL 结果（跨轮状态）；
    - agent：多跳/组合问题走手写 ReAct 循环（query_sql / retrieve_vector / read_memory 三工具），
      Agent 未收敛或输出不可解析时回退快路径（_run_route），保证异常不拖垮问答；
    - session_id 非 None 时启用会话记忆（历史 + 上轮结果暂存）；为 None 保持 V10 无记忆行为；
    - 依赖注入（测试用）与无 session_id：不启用记忆，行为与 V10.3 完全一致（测试零影响）。
    """
    injected = not (llm is None and retriever is None and db is None)
    q = question.strip()
    session = _session(session_id) if session_id is not None else None
    if not injected:
        hit = _answer_cache.get(q)
        if hit is not None:
            print(f"  [缓存] 命中相同问题（{len(_answer_cache)} 条缓存中），直接返回")
            if session is not None:
                _record_session(session, q, hit)  # 命中也要记账，追问指代才接得上
            return hit
    if llm is None and retriever is None and db is None:
        _llm, _db = globals()["llm"], globals()["db"]
        _chain, _router = chain, router_chain
    else:
        _llm = llm or globals()["llm"]
        _retriever = retriever or globals()["retriever"]
        _db = db or globals()["db"]
        _chain = build_qa_chain(_llm, _retriever)
        _router = router_prompt | _llm | StrOutputParser()

    # 历史只喂给路由器判断"是否引用上轮"，不直接参与 vector/sql 链的回答
    history_hint = ""
    if session is not None:
        history_hint = "\n".join(
            f"{role}：{text}" for role, text in session["history"][-MAX_HISTORY_ROUNDS * 2:]
        ) or "（无历史）"

    route = _router.invoke({"question": q, "history_hint": history_hint}).strip()
    print(f"  [路由器] 问题类型：{route}")
    if route == "agent":
        response = _run_agent(q, _llm, _db, _chain, session)
        if response is None:  # Agent 未收敛/输出不可解析 → 回退快路径（路由器再判一次）
            print("  [路由器] Agent 未收敛，回退快路径")
            response = _run_route(q, _llm, _db, _chain, _router, history_hint=history_hint)
    elif route == "memory" and session is not None:
        response = _answer_with_memory(q, session, _llm, _db, _chain, _router)
    elif route == "sql":
        ans = sql_answer(q, _llm, _db)
        if ans is not None:
            response = ans
        else:
            print("  [路由器] SQL 链无结果，降级向量检索")
            response = _chain.invoke({"question": q}).content
    else:
        response = _chain.invoke({"question": q}).content

    if session is not None:
        _record_session(session, q, response)
    if not injected:
        if len(_answer_cache) >= _MAX_CACHE_SIZE:
            _answer_cache.pop(next(iter(_answer_cache)))  # dict 保序，超限弹出最旧
        _answer_cache[q] = response
    return response


# ========== 7. 运行 ==========
if __name__ == "__main__":
    init()  # CLI 入口显式初始化（import 不再自动构建索引）
    print("多轮对话模式（第 9 步多轮记忆）：输入 exit / quit / 退出 结束")
    print("可连续追问并用指代，如先问\"李云龙的妻子是谁？\"再问\"那她们结局如何？\"\n")
    session_id = "cli"  # 固定会话：整个 CLI 进程内共享历史与上轮结果
    while True:
        try:
            question = input("你：")
        except (EOFError, KeyboardInterrupt):
            print("\n再见")
            break
        q = question.strip()
        if not q:
            continue
        if q.lower() in ("exit", "quit", "退出"):
            break
        response = ask(q, session_id=session_id)
        print(f"回答：{response}\n")

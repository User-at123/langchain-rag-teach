"""Text-to-SQL 数据层（第 5 步，V9.0 新增）。

把 docs/ 下的 .xlsx 导入 SQLite（零新增依赖：sqlite3 是标准库，openpyxl 已存在），
提供三件事：
1. 建库导入：表名 = sheet 名，列名 = 表头（第 1 行），与 load_xlsx 同源；
   幂等：表已建且有数据则跳过，重复运行不重复导入。
2. 只读执行器：只允许单条 SELECT（防注入）+ 中文标识符必须是真实列名
   （中文不可能是 SQL 关键字，所以对中文表头这条校验 100% 可靠，
   防止 LLM 捏造"诞生时间"这种不存在的列）。
3. schema 文本：动态生成"表 + 列 + 样例数据"，供 Text-to-SQL 提示词使用。
   列名不用手动维护——术语翻译交给 LLM（它的世界常识），
   这里只保证"给 LLM 看的东西永远是数据库里的真实情况"。
"""

import os
import re
import sqlite3
import threading

from openpyxl import load_workbook

DOCS_DIR = os.getenv("DOCS_DIR", "./docs")
DB_PATH = os.getenv("DB_PATH", "./data.db")


class SQLiteDb:
    def __init__(self, db_path=DB_PATH, docs_dir=DOCS_DIR):
        # check_same_thread=False：FastAPI 的同步接口跑在线程池里，
        # 请求线程和创建连接的主线程不是同一个，必须允许跨线程使用；
        # 并发安全靠 self._lock 保证（同一时刻只有一个线程访问连接）。
        self.conn = sqlite3.connect(db_path, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self._lock = threading.Lock()
        self._import_xlsx_files(docs_dir)
        print(f"SQLite 就绪：{db_path}（表：{self._tables()}）")

    # ========== 建库导入 ==========

    def _import_xlsx_files(self, docs_dir):
        if not os.path.isdir(docs_dir):
            return
        for root, _, files in os.walk(docs_dir):
            for name in files:
                if name.lower().endswith(".xlsx"):
                    self._import_xlsx(os.path.join(root, name))

    def _import_xlsx(self, path):
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.worksheets:
            table = sheet.title
            if self._has_data(table):
                continue  # 幂等：已导入过则跳过
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            header_raw = ["" if c is None else str(c).strip() for c in rows[0]]
            if not any(header_raw):
                continue  # 无表头的表跳过（与 load_xlsx 行为一致）
            # 空表头列用占位名，保证列位与数据行对齐
            header = [h if h else f"列{i + 1}" for i, h in enumerate(header_raw)]
            cols_sql = ", ".join(f'"{h}" TEXT' for h in header)
            self.conn.execute(f'CREATE TABLE IF NOT EXISTS "{table}" ({cols_sql})')
            data = []
            for row in rows[1:]:
                cells = ["" if c is None else str(c).strip() for c in row]
                if not any(cells):
                    continue
                if cells == header_raw:
                    continue  # 跳过重复打印的表头行
                cells = (cells + [""] * len(header))[: len(header)]  # 补齐/截断列数
                data.append(tuple(cells))
            if data:
                ph = ", ".join("?" for _ in header)
                self.conn.executemany(f'INSERT INTO "{table}" VALUES ({ph})', data)
        wb.close()
        self.conn.commit()

    def _tables(self):
        rows = self.conn.execute(
            "SELECT name FROM sqlite_master"
            " WHERE type='table' AND name NOT LIKE 'sqlite_%'"
        ).fetchall()
        return [r[0] for r in rows]

    def _columns(self, table):
        return [r[1] for r in self.conn.execute(f'PRAGMA table_info("{table}")')]

    def _has_data(self, table):
        if table not in self._tables():
            return False
        return self.conn.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0] > 0

    # ========== 只读执行器 ==========

    def query(self, sql):
        """执行只读 SELECT。三道校验：单条语句 / SELECT 开头 / 中文标识符合法。"""
        stmt = sql.strip().rstrip(";").strip()
        if not stmt or ";" in stmt:
            raise ValueError("只允许单条 SELECT 语句")
        if not stmt.upper().startswith("SELECT"):
            raise ValueError("只允许 SELECT 查询（只读白名单）")
        self._validate_identifiers(stmt)
        with self._lock:
            cur = self.conn.execute(stmt)
            cols = [d[0] for d in cur.description]
            rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        return cols, rows

    def _validate_identifiers(self, sql):
        """SQL 中的中文标识符必须属于某张表的表名或列名。
        先把单引号字符串字面量剥掉（字符串里的中文是值，不是标识符）；
        AS 别名是临时标识符（如 COUNT(*) AS 客户数量），放行进合法集合；
        剩下的中文 token 若不在合法集合里，说明 LLM 捏造了列名 → 拦截。"""
        legal = set(self._tables())
        for t in self._tables():
            legal.update(self._columns(t))
        sql_no_str = re.sub(r"'[^']*'", "", sql)
        for alias in re.findall(r"AS\s+([\u4e00-\u9fff]+)", sql_no_str):
            legal.add(alias)  # 别名不是捏造列名，放行（含 ORDER BY 别名引用）
        for tok in re.findall(r"[\u4e00-\u9fff]+", sql_no_str):
            if tok not in legal:
                raise ValueError(
                    f"标识符 '{tok}' 不在任何表中（合法：{sorted(legal)}）"
                )

    # ========== schema 文本（给 Text-to-SQL 提示词） ==========

    def schema_text(self, sample_rows=2):
        """动态生成表结构描述：每个表的列名 + 前几行样例数据。"""
        parts = []
        with self._lock:
            for t in self._tables():
                cols = self._columns(t)
                sample = self.conn.execute(
                    f'SELECT * FROM "{t}" LIMIT {sample_rows}'
                ).fetchall()
                parts.append(f'表 "{t}"，列: {", ".join(cols)}')
                for row in sample:
                    parts.append(
                        "  样例: " + " | ".join(f"{c}: {v}" for c, v in zip(cols, row))
                    )
        return "\n".join(parts)
